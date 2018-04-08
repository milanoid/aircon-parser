from openpyxl import load_workbook


class OutputWriter:

    def __init__(self, passengers, header, filename="output.xlsx"):
        self.passengers = passengers
        self.header = header
        self.filename = filename
        self.destination_workbook = load_workbook(self.filename)

    def write_passengers(self):
        i = 5  # first passenger starts on row 5
        passenger_number = 1
        for passenger in self.passengers:
            self.destination_workbook.active.cell(i, 1).value = passenger_number
            self.destination_workbook.active.cell(i, 2).value = passenger.gender
            self.destination_workbook.active.cell(i, 3).value = passenger.last_name
            self.destination_workbook.active.cell(i, 4).value = passenger.first_name

            if passenger.is_infant:
                self.destination_workbook.active.cell(i, 5).value = "Yes"

            i += 1
            passenger_number += 1

        return self

    def write_header(self):
        self.destination_workbook.active.cell(2, 2).value = self.header.routing
        self.destination_workbook.active.cell(2, 3).value = self.header.flight_date
        self.destination_workbook.active.cell(2, 4).value = self.header.flight_number

        # e.g. 183 (adults) + 1 (infant)
        infants = 0
        adults = 0
        for passenger in self.passengers:
            if passenger.is_infant:
                infants += 1
            else:
                adults += 1

        self.destination_workbook.active.cell(2, 6).value = str(adults) + "+" + str(infants)

        return self

    def save(self):
        self.destination_workbook.save(self.filename)
