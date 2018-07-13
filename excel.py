import glob
from openpyxl import load_workbook
from subprocess import call
from AirconParser import PassengerParser, HeaderParser, InfantParser
from OutputWriter import OutputWriter
from Passenger import Passenger


def main():
    convert_xls_to_xlsx()
    excel_files_to_process = glob.glob('data\LISTADO_SITA*.xlsx')

    for file in excel_files_to_process:
        convert_to_aircon_format(file)


def convert_xls_to_xlsx():
    call(["convert_xls_to_xlsx.cmd"])


def convert_to_aircon_format(source_workbook_filename):
    workbook_source = load_workbook(source_workbook_filename)
    sheet = workbook_source.active
    passengers = []

    for i in range(8, sheet.max_row):

        if sheet.cell(i, 3).value is not None:
            passenger = PassengerParser(sheet.cell(i, 3).value)
            passengers.append(Passenger(passenger.first_name, passenger.last_name, passenger.gender))
            if passenger.has_infant:
                infant = InfantParser(sheet.cell(i, 4).value)
                passengers.append(
                    Passenger(
                        first_name=infant.first_name,
                        last_name=infant.last_name,
                        gender="?",
                        has_infant=False,
                        is_infant=True
                    ))

    header_to_parse = sheet.cell(3, 1).value

    writer = OutputWriter(
        passengers=passengers,
        header=HeaderParser(header_to_parse),
        filename=source_workbook_filename + "aircon.xlsx"
    )
    writer.write_passengers().write_header().save()


if __name__ == "__main__":
    main()
